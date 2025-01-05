from __future__ import annotations

import logging
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, ClassVar, Dict, Optional, Set

import docx2txt
import fitz  # PyMuPDF
import nltk
from bs4 import BeautifulSoup
from pandas import read_excel
from pdfminer.high_level import extract_text as extract_text_pdf
from PIL import Image
import pytesseract
from openpyxl import load_workbook

from .exceptions import DocumentProcessingException
from .file_system import FileInfo

logger = logging.getLogger(__name__)

@dataclass
class ProcessedDocument:
    """Represents a processed document with extracted content and metadata."""
    file_info: FileInfo
    content: str
    metadata: dict[str, Any]
    summary: str
    keywords: set[str]
    language: str
    processing_time: float
    processed_date: datetime
    error: Optional[str] = None

class DocumentProcessor:
    """Process and extract content from various document types."""
    
    # Supported file extensions mapped to their processing methods
    SUPPORTED_EXTENSIONS: ClassVar[Dict[str, str]] = {
        '.txt': '_process_text',
        '.pdf': '_process_pdf',
        '.docx': '_process_docx',
        '.doc': '_process_doc',
        '.xlsx': '_process_excel',
        '.xls': '_process_excel',
        '.csv': '_process_csv',
        '.html': '_process_html',
        '.htm': '_process_html',
        '.jpg': '_process_image',
        '.jpeg': '_process_image',
        '.png': '_process_image',
        '.tiff': '_process_image'
    }

    def __init__(self, nltk_data_path: Optional[str] = None) -> None:
        """Initialize document processor.
        
        Args:
            nltk_data_path: Optional path to NLTK data directory
        """
        self._setup_nltk(nltk_data_path)
        self._initialize_tesseract()
        
    def _setup_nltk(self, data_path: Optional[str]) -> None:
        """Setup NLTK resources."""
        try:
            if data_path:
                nltk.data.path.append(data_path)
            
            required_packages = [
                'punkt',
                'averaged_perceptron_tagger',
                'maxent_ne_chunker',
                'words',
                'stopwords'
            ]
            
            for package in required_packages:
                try:
                    nltk.data.find(f'tokenizers/{package}')
                except LookupError:
                    nltk.download(package, quiet=True)
                    
        except Exception as e:
            logger.error(f"Error setting up NLTK: {e}")
            raise DocumentProcessingException(f"Failed to initialize NLTK: {e}")

    def _initialize_tesseract(self) -> None:
        """Initialize Tesseract OCR."""
        try:
            pytesseract.get_tesseract_version()
        except Exception as e:
            logger.error(f"Error initializing Tesseract: {e}")
            raise DocumentProcessingException(
                "Tesseract OCR is not properly installed. Please install Tesseract."
            )

    def process_document(self, file_info: FileInfo) -> ProcessedDocument:
        """Process a document and extract its content.
        
        Args:
            file_info: FileInfo object containing file information
            
        Returns:
            ProcessedDocument object with extracted content and metadata
        """
        start_time = datetime.now()
        error = None
        content = ""
        metadata: dict[str, Any] = {}

        try:
            extension = file_info.extension.lower()
            if extension not in self.SUPPORTED_EXTENSIONS:
                raise DocumentProcessingException(
                    f"Unsupported file type: {extension}"
                )

            # Get processing method
            process_method = getattr(
                self,
                self.SUPPORTED_EXTENSIONS[extension]
            )
            
            # Process document
            content, metadata = process_method(file_info.path)
            
            # Extract additional information
            summary = self._generate_summary(content)
            keywords = self._extract_keywords(content)
            language = self._detect_language(content)

        except Exception as e:
            logger.error(f"Error processing document {file_info.path}: {e}")
            error = str(e)
            summary = ""
            keywords = set()
            language = ""

        processing_time = (datetime.now() - start_time).total_seconds()
        
        return ProcessedDocument(
            file_info=file_info,
            content=content,
            metadata=metadata,
            summary=summary,
            keywords=keywords,
            language=language,
            processing_time=processing_time,
            processed_date=datetime.now(),
            error=error
        )

    def _process_text(self, file_path: Path) -> tuple[str, dict]:
        """Process plain text files."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            metadata = {
                'encoding': 'utf-8',
                'type': 'text'
            }
            return content, metadata
        except UnicodeDecodeError:
            # Try different encodings
            encodings = ['latin-1', 'cp1252', 'iso-8859-1']
            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        content = f.read()
                    metadata = {
                        'encoding': encoding,
                        'type': 'text'
                    }
                    return content, metadata
                except UnicodeDecodeError:
                    continue
            raise DocumentProcessingException("Failed to decode text file")

    def _process_pdf(self, file_path: Path) -> tuple[str, dict]:
        """Process PDF files using both PyMuPDF and pdfminer for robustness."""
        content = ""
        metadata = {}
        
        # Try PyMuPDF first
        try:
            doc = fitz.open(file_path)
            try:
                content = " ".join(page.get_text() for page in doc)
                metadata = {
                    'page_count': len(doc),
                    'author': doc.metadata.get('author', ''),
                    'title': doc.metadata.get('title', ''),
                    'subject': doc.metadata.get('subject', ''),
                    'keywords': doc.metadata.get('keywords', ''),
                    'creator': doc.metadata.get('creator', ''),
                    'producer': doc.metadata.get('producer', ''),
                    'type': 'pdf'
                }
            finally:
                doc.close()
        except Exception as e:
            logger.warning(f"PyMuPDF failed, trying pdfminer: {e}")
            
            # Fallback to pdfminer
            try:
                content = extract_text_pdf(file_path)
                metadata = {
                    'type': 'pdf',
                    'extraction_method': 'pdfminer'
                }
            except Exception as pdf_error:
                raise DocumentProcessingException(
                    f"Failed to process PDF with both methods: {pdf_error}"
                )
                
        return content, metadata

    def _process_docx(self, file_path: Path) -> tuple[str, dict]:
        """Process DOCX files."""
        try:
            content = docx2txt.process(file_path)
            metadata = {
                'type': 'docx'
            }
            return content, metadata
        except Exception as e:
            raise DocumentProcessingException(f"Failed to process DOCX: {e}")

    def _process_doc(self, file_path: Path) -> tuple[str, dict]:
        """Process legacy DOC files."""
        import subprocess
        import platform
        
        if platform.system() == 'Windows':
            try:
                import win32com.client
                word = win32com.client.Dispatch("Word.Application")
                word.visible = False
                try:
                    doc = word.Documents.Open(str(file_path))
                    content = doc.Content.Text
                    metadata = {
                        'type': 'doc',
                        'author': doc.Author,
                        'creation_date': doc.CreationDate,
                        'last_modified': doc.LastSavedTime
                    }
                    return content, metadata
                finally:
                    word.Quit()
            except Exception:
                logger.warning("COM automation failed, trying LibreOffice")
        
        # Fallback to LibreOffice conversion
        try:
            with tempfile.NamedTemporaryFile(suffix='.txt') as tmp:
                subprocess.run(
                    ['soffice', '--headless', '--convert-to', 'txt:Text', 
                     str(file_path), '--outdir', tmp.name],
                    capture_output=True,
                    check=True
                )
                txt_path = Path(tmp.name).with_suffix('.txt')
                with open(txt_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                metadata = {
                    'type': 'doc',
                    'conversion_method': 'libreoffice'
                }
                return content, metadata
        except Exception as e:
            raise DocumentProcessingException(f"Failed to process DOC: {e}")

    def _process_excel(self, file_path: Path) -> tuple[str, dict]:
        """Process Excel files."""
        try:
            # First try pandas for basic content
            df = read_excel(file_path)
            content = df.to_string()
            
            # Then use openpyxl for metadata
            wb = load_workbook(filename=file_path, read_only=True)
            metadata = {
                'type': 'excel',
                'sheet_names': wb.sheetnames,
                'sheet_count': len(wb.sheetnames),
                'properties': {
                    'creator': wb.properties.creator,
                    'created': wb.properties.created,
                    'modified': wb.properties.modified,
                    'title': wb.properties.title,
                    'subject': wb.properties.subject,
                    'keywords': wb.properties.keywords,
                    'category': wb.properties.category,
                }
            }
            wb.close()
            
            return content, metadata
        except Exception as e:
            raise DocumentProcessingException(f"Failed to process Excel file: {e}")

    def _process_csv(self, file_path: Path) -> tuple[str, dict]:
        """Process CSV files."""
        import csv
        try:
            content_lines = []
            with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
                reader = csv.reader(csvfile)
                for row in reader:
                    content_lines.append(','.join(row))
            
            content = '\n'.join(content_lines)
            metadata = {
                'type': 'csv',
                'row_count': len(content_lines)
            }
            return content, metadata
        except Exception as e:
            raise DocumentProcessingException(f"Failed to process CSV: {e}")

    def _process_html(self, file_path: Path) -> tuple[str, dict]:
        """Process HTML files."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                soup = BeautifulSoup(f.read(), 'html.parser')
                
            # Remove script and style elements
            for script in soup(["script", "style"]):
                script.decompose()
                
            # Extract text content
            content = soup.get_text(separator=' ', strip=True)
            
            # Extract metadata
            metadata = {
                'type': 'html',
                'title': soup.title.string if soup.title else '',
                'meta_tags': {
                    meta.get('name', meta.get('property', '')): meta.get('content', '')
                    for meta in soup.find_all('meta')
                    if meta.get('name') or meta.get('property')
                }
            }
            
            return content, metadata
        except Exception as e:
            raise DocumentProcessingException(f"Failed to process HTML: {e}")

    def _process_image(self, file_path: Path) -> tuple[str, dict]:
        """Process images using OCR."""
        try:
            image = Image.open(file_path)
            
            # Extract image metadata
            metadata = {
                'type': 'image',
                'format': image.format,
                'mode': image.mode,
                'size': image.size,
                'width': image.width,
                'height': image.height
            }
            
            # Perform OCR
            content = pytesseract.image_to_string(image)
            
            return content, metadata
        except Exception as e:
            raise DocumentProcessingException(f"Failed to process image: {e}")

    def _generate_summary(self, text: str, max_sentences: int = 3) -> str:
        """Generate a summary of the text content."""
        try:
            if not text.strip():
                return ""
            
            sentences = nltk.sent_tokenize(text)
            if len(sentences) <= max_sentences:
                return text
                
            # Simple extractive summarization
            word_freq = {}
            for word in nltk.word_tokenize(text.lower()):
                if word.isalnum():
                    word_freq[word] = word_freq.get(word, 0) + 1
                    
            sentence_scores = {}
            for sentence in sentences:
                for word in nltk.word_tokenize(sentence.lower()):
                    if word in word_freq:
                        sentence_scores[sentence] = sentence_scores.get(sentence, 0) + word_freq[word]
                        
            summary_sentences = sorted(
                sentence_scores.items(),
                key=lambda x: x[1],
                reverse=True
            )[:max_sentences]
            
            summary = ' '.join(sent[0] for sent in sorted(
                summary_sentences,
                key=lambda x: sentences.index(x[0])
            ))
            
            return summary
            
        except Exception as e:
            logger.error(f"Error generating summary: {e}")
            return ""

    def _extract_keywords(self, text: str, max_keywords: int = 10) -> Set[str]:
        """Extract key phrases and entities from text."""
        try:
            if not text.strip():
                return set()
                
            # Get stopwords
            stop_words = set(nltk.corpus.stopwords.words('english'))
            
            # Tokenize and tag parts of speech
            tokens = nltk.word_tokenize(text.lower())
            pos_tags = nltk.pos_tag(tokens)
            
            # Extract noun phrases
            grammar = "NP: {<JJ>*<NN.*>+}"
            chunk_parser = nltk.RegexpParser(grammar)
            tree = chunk_parser.parse(pos_tags)
            
            noun_phrases = set()
            for subtree in tree.subtrees(filter=lambda t: t.label() == 'NP'):
                phrase = ' '.join(word for word, tag in subtree.leaves())
                if len(phrase.split()) <= 3:  # Limit phrase length
                    noun_phrases.add(phrase)
            
            # Extract named entities
            named_entities = set()
            for chunk in nltk.ne_chunk(pos_tags):
                if hasattr(chunk, 'label'):
                    entity = ' '.join(c[0] for c in chunk.leaves())
                    named_entities.add(entity)
            
            # Get most common words (excluding stopwords)
            word_freq = {}
            for word, tag in pos_tags:
                if (word.isalnum() and word not in stop_words and 
                    tag.startswith(('NN', 'VB', 'JJ'))):
                    word_freq[word] = word_freq.get(word, 0) + 1
            
            common_words = set(sorted(
                word_freq.items(),
                key=lambda x: x[1],
                reverse=True
            )[:max_keywords])
            
            # Combine all keywords
            keywords = noun_phrases | named_entities | {word for word, _ in common_words}
            
            # Return top keywords
            return set(sorted(keywords, key=len, reverse=True)[:max_keywords])
            
        except Exception as e:
            logger.error(f"Error extracting keywords: {e}")
            return set()

    def _detect_language(self, text: str) -> str:
        """Detect the language of the text."""
        try:
            if not text.strip():
                return ""
                
            from langdetect import detect, DetectorFactory
            DetectorFactory.seed = 0  # For consistent results
            
            return detect(text)
        except Exception as e:
            logger.error(f"Error detecting language: {e}")
            return ""

    def _extract_text_chunks(
        self,
        text: str,
        chunk_size: int = 1000,
        overlap: int = 100
    ) -> list[str]:
        """Split text into overlapping chunks for processing."""
        if not text:
            return []
            
        chunks = []
        sentences = nltk.sent_tokenize(text)
        current_chunk = []
        current_size = 0
        
        for sentence in sentences:
            sentence_size = len(sentence)
            
            if current_size + sentence_size > chunk_size and current_chunk:
                # Add current chunk to results
                chunks.append(' '.join(current_chunk))
                
                # Keep last few sentences for overlap
                overlap_size = 0
                overlap_sentences = []
                
                for s in reversed(current_chunk):
                    overlap_size += len(s)
                    if overlap_size > overlap:
                        break
                    overlap_sentences.append(s)
                
                current_chunk = list(reversed(overlap_sentences))
                current_size = sum(len(s) for s in current_chunk)
            
            current_chunk.append(sentence)
            current_size += sentence_size
        
        if current_chunk:
            chunks.append(' '.join(current_chunk))
        
        return chunks